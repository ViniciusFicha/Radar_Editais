import requests
import json
import os
from datetime import datetime
import openpyxl
import feedparser
import smtplib
from email.mime.text import MIMEText

# =========================
# CARREGAR CONFIG
# =========================
with open("config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

ARQUIVO_EXCEL = config["excel"]["arquivo"]
PALAVRAS_CHAVE = config["filtros"]["palavras_chave"]

EMAIL_CONFIG = config.get("email", {})
EMAIL_ATIVAR = EMAIL_CONFIG.get("ativar", False)
EMAIL_REMETENTE = EMAIL_CONFIG.get("remetente", "")
EMAIL_SENHA = EMAIL_CONFIG.get("senha_app", "")
EMAIL_DESTINATARIOS = EMAIL_CONFIG.get(
    "destinatarios",
    ["guilherme.amorim@rnsj.com.br"]  #  diretor já incluído
)

# =========================
# CRIAR EXCEL
# =========================
if not os.path.exists(ARQUIVO_EXCEL):
    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        "Data", "Título", "Órgão", "UF", "Link",
        "Aderência", "Decisão", "Status",
        "Resumo Executivo", "Requisitos", "Oportunidades", "Riscos"
    ])

    wb.save(ARQUIVO_EXCEL)

# =========================
# ABRIR EXCEL
# =========================
wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
ws = wb.active

links_existentes = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[4]:
        links_existentes.add(row[4])

# =========================
# IA SIMULADA
# =========================
def analisar_edital(titulo):
    t = titulo.lower()

    resumo = f"Oportunidade relacionada a: {titulo}"

    if "experiência" in t or "exig" in t:
        requisitos = "Possível exigência de experiência prévia ou qualificação técnica"
    else:
        requisitos = "Requisitos não detalhados na fonte"

    if "consultoria" in t or "jurídica" in t:
        oportunidades = "Possibilidade de atuação jurídica consultiva"
    elif "projeto" in t or "pesquisa" in t:
        oportunidades = "Espaço para produção técnica ou acadêmica"
    else:
        oportunidades = "Oportunidade geral a avaliar"

    score = 0
    if "consultoria" in t: score += 10
    if "jurídica" in t or "juridica" in t: score += 10
    if "advocacia" in t: score += 8
    if "licitação" in t: score += 5
    if "contratação" in t: score += 5
    if "projeto" in t: score += 3

    if score >= 15:
        aderencia = "ALTA"
    elif score >= 7:
        aderencia = "MÉDIA"
    else:
        aderencia = "BAIXA"

    if aderencia == "ALTA":
        decisao = "PARTICIPAR"
    elif aderencia == "MÉDIA":
        decisao = "MONITORAR"
    else:
        decisao = "DESCARTAR"

    if "temporário" in t:
        riscos = "Prazo curto ou contrato limitado"
    elif "complexo" in t:
        riscos = "Alta complexidade operacional"
    else:
        riscos = "Baixos ou não identificados"

    return resumo, requisitos, oportunidades, riscos, aderencia, decisao

# =========================
# BUSCAR RSS
# =========================
url = "https://g1.globo.com/rss/g1/concursos-e-emprego/"
feed = feedparser.parse(url)

print("Itens encontrados:", len(feed.entries))

editais = []

for entry in feed.entries:
    titulo = entry.title
    link = entry.link
    titulo_lower = titulo.lower()

    if not any(p.lower() in titulo_lower for p in PALAVRAS_CHAVE):
        if "concurso" not in titulo_lower and "processo seletivo" not in titulo_lower:
            continue

    if link in links_existentes:
        continue

    editais.append({
        "titulo": titulo,
        "orgao": "Fonte G1",
        "uf": "BR",
        "link": link
    })

# =========================
# INSERIR DADOS
# =========================
novos = 0

for e in editais:
    print("Analisando:", e["titulo"])

    resumo, requisitos, oportunidades, riscos, aderencia, decisao = analisar_edital(e["titulo"])

    ws.append([
        datetime.now().strftime("%Y-%m-%d"),
        e["titulo"],
        e["orgao"],
        e["uf"],
        e["link"],
        aderencia,
        decisao,
        "Novo",
        resumo,
        requisitos,
        oportunidades,
        riscos
    ])

    novos += 1

# =========================
# ENVIAR EMAIL
# =========================
def enviar_email(editais):
    if not EMAIL_ATIVAR or not editais:
        return

    corpo = "RADAR DE EDITAIS - NOVAS OPORTUNIDADES\n\n"

    for e in editais:
        corpo += f"Título: {e['titulo']}\n"
        corpo += f"Órgão: {e['orgao']}\n"
        corpo += f"Link: {e['link']}\n"
        corpo += "---------------------------\n"

    msg = MIMEText(corpo)
    msg["Subject"] = "Radar de Editais - Atualização Automática"
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = ", ".join(EMAIL_DESTINATARIOS)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_REMETENTE, EMAIL_SENHA)
            server.sendmail(EMAIL_REMETENTE, EMAIL_DESTINATARIOS, msg.as_string())

        print("Email enviado com sucesso.")

    except Exception as e:
        print("Erro ao enviar email:", e)

# ✅ CHAMA O EMAIL
enviar_email(editais)

# =========================
# SALVAR
# =========================
wb.save(ARQUIVO_EXCEL)

print(f"{novos} novos editais analisados.")