import json
import os
from datetime import datetime
import openpyxl
import requests
from bs4 import BeautifulSoup

# =========================
# CONFIG
# =========================
with open("config.json", "r", encoding="utf-8") as f:
    config = json.load(f)

ARQUIVO_EXCEL = config["excel"]["arquivo"]
PALAVRAS_CHAVE = config["filtros"]["palavras_chave"]

# =========================
# CRIAR EXCEL
# =========================
if not os.path.exists(ARQUIVO_EXCEL):
    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append([
        "Data", "Título", "Órgão", "UF", "Link",
        "Aderência", "Decisão", "Status",
        "Resumo Executivo", "Requisitos", "Oportunidades", "Riscos"
    ])

    wb.save(ARQUIVO_EXCEL)

# =========================
# ABRIR EXCEL
# =========================
try:
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
except PermissionError:
    print("ERRO: Feche o Excel antes de rodar o sistema.")
    exit()

ws = wb.active

links_existentes = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[4]:
        links_existentes.add(row[4])

# =========================
# IA SIMULADA
# =========================
def analisar_edital(titulo):
    t = titulo.lower()

    resumo = f"Oportunidade relacionada a '{titulo}'"

    requisitos = "Não identificados"
    if any(x in t for x in ["experience", "experiência", "requirement"]):
        requisitos = "Possível exigência técnica"

    oportunidades = "Avaliar aderência"
    if "consult" in t:
        oportunidades = "Atuação consultiva"
    elif "project" in t or "projeto" in t:
        oportunidades = "Produção técnica"

    score = 0
    for p in ["consult", "governance", "policy", "legal", "advisor", "compliance"]:
        if p in t:
            score += 5

    if score >= 10:
        aderencia = "ALTA"
        decisao = "PARTICIPAR"
    elif score >= 5:
        aderencia = "MÉDIA"
        decisao = "MONITORAR"
    else:
        aderencia = "BAIXA"
        decisao = "DESCARTAR"

    riscos = "Baixos"

    return resumo, requisitos, oportunidades, riscos, aderencia, decisao

# =========================
# FILTROS
# =========================
PALAVRAS_LIXO = ["privacy", "cookie", "terms", "about", "contact"]

PALAVRAS_FORTES = [
    "consult", "consultant",
    "policy", "governance",
    "legal", "law",
    "advisor", "compliance"
]

# =========================
# BUSCAR UN JOBS
# =========================
def buscar_un_jobs():
    url = "https://unjobs.org"

    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        resultados = []

        for a in soup.select("a")[:200]:
            titulo = a.get_text(strip=True)
            href = a.get("href")

            if not titulo or not href:
                continue

            t = titulo.lower()

            # remover lixo
            if any(x in t for x in PALAVRAS_LIXO):
                continue

            # filtro principal
            if not any(p in t for p in PALAVRAS_FORTES):
                continue

            if not href.startswith("http"):
                href = "https://unjobs.org" + href

            if href in links_existentes:
                continue

            resultados.append({
                "titulo": titulo,
                "orgao": "UN Jobs",
                "uf": "INT",
                "link": href
            })

        print("UN Jobs filtrados:", len(resultados))
        return resultados[:10]

    except Exception as e:
        print("Erro UN Jobs:", e)
        return []

# =========================
# BUSCAR RELIEFWEB
# =========================
def buscar_reliefweb():
    url = "https://reliefweb.int/jobs"

    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        resultados = []

        for a in soup.select("a")[:200]:
            titulo = a.get_text(strip=True)
            href = a.get("href")

            if not titulo or not href:
                continue

            t = titulo.lower()

            if any(x in t for x in PALAVRAS_LIXO):
                continue

            if not any(p in t for p in PALAVRAS_FORTES):
                continue

            if not href.startswith("http"):
                href = "https://reliefweb.int" + href

            if href in links_existentes:
                continue

            resultados.append({
                "titulo": titulo,
                "orgao": "ReliefWeb",
                "uf": "INT",
                "link": href
            })

        print("ReliefWeb filtrados:", len(resultados))
        return resultados[:10]

    except Exception as e:
        print("Erro ReliefWeb:", e)
        return []

# =========================
# COLETAR
# =========================
editais = []

for fonte in [buscar_un_jobs(), buscar_reliefweb()]:
    for e in fonte:
        if e["link"] not in links_existentes:
            editais.append(e)

# fallback
if not editais:
    print("Nenhuma oportunidade encontrada - fallback ativado")

    editais.append({
        "titulo": "Consultant – Public Policy",
        "orgao": "Referência",
        "uf": "INT",
        "link": "https://unjobs.org"
    })

# =========================
# INSERIR
# =========================
novos = 0

for e in editais:
    print("Analisando:", e["titulo"])

    resumo, requisitos, oportunidades, riscos, aderencia, decisao = analisar_edital(e["titulo"])

    ws.append([
        datetime.now().strftime("%Y-%m-%d"),
        e["titulo"], e["orgao"], e["uf"], e["link"],
        aderencia, decisao, "Novo",
        resumo, requisitos, oportunidades, riscos
    ])

    novos += 1

# =========================
# SALVAR
# =========================
try:
    wb.save(ARQUIVO_EXCEL)
except PermissionError:
    print("ERRO: Feche o Excel antes de salvar.")

print(f"{novos} oportunidades estratégicas registradas.")